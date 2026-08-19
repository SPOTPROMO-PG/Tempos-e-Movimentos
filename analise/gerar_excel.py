# -*- coding: utf-8 -*-
"""
Monta o Excel entregue ao cliente a partir das saídas de jornada.py.

A aba "Movimentos" é a base: guarda os minutos equivalentes de cada
movimento em cada canal. Todas as outras abas de número saem dela por
SUMIFS, então o cliente consegue auditar de onde veio cada célula e o
arquivo recalcula sozinho se alguma linha for ajustada.
"""

import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

mov = pd.read_csv('analise/saida/jornada_por_canal.csv')
vis = pd.read_csv('analise/saida/resumo_visitas.csv')
cob = pd.read_csv('analise/saida/cobertura.csv')
meta = json.load(open('analise/saida/meta.json', encoding='utf-8'))
meta['periodo'] = ['{2}/{1}/{0}'.format(*d.split('-')) for d in meta['periodo']]   # ISO -> dd/mm/aaaa
desl = json.load(open('analise/saida/desloc.json', encoding='utf-8'))
extra = json.load(open('analise/saida/extra.json', encoding='utf-8'))

TINTA = '1C5570'
F = 'Arial'
h1 = Font(name=F, size=16, bold=True, color=TINTA)
h2 = Font(name=F, size=11, bold=True, color=TINTA)
hd = Font(name=F, size=9, bold=True, color='FFFFFF')
nm = Font(name=F, size=10)
bd = Font(name=F, size=10, bold=True)
it = Font(name=F, size=9, italic=True, color='6D7887')
fill = PatternFill('solid', fgColor=TINTA)
zebra = PatternFill('solid', fgColor='F2F5F8')
thin = Border(bottom=Side('thin', color='D8DEE6'))

BLOCOS = ['Entrada', 'Abertura', 'Estoque', 'Check Out', 'Ponto Natural',
          'Ponto Extra', 'Outras Atividades', 'Saída']
CLASSES = ['Execução (agrega valor)', 'Apoio e análise',
           'Deslocamento e busca', 'Espera e burocracia']
POUCAS = {'CLUB', 'LASA', 'PERFUMARIA'}   # amostra reduzida


def cabecalho(ws, row, cols, larguras):
    for i, (c, w) in enumerate(zip(cols, larguras), 1):
        cel = ws.cell(row=row, column=i, value=c)
        cel.font = hd
        cel.fill = fill
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def zebrar(ws, R, ncols, i):
    if i % 2:
        for c in range(1, ncols + 1):
            ws.cell(row=R, column=c).fill = zebra


wb = Workbook()
wb.remove(wb.active)
nvis = vis.groupby('canal').loja.size().to_dict()
ordc = vis.groupby('canal').size().sort_values(ascending=False).index.tolist()

# ---------------------------------------------------------------- Leia-me
ws = wb.create_sheet('Leia-me')
ws.column_dimensions['A'].width = 112
LINHAS = [
    ('Estudo de Tempos e Movimentos em Loja', 't'),
    ('Promotores de merchandising — P&G / SPOT', 's'),
    ('Coleta de {} a {}'.format(*meta['periodo']), 's'),
    ('', ''),
    ('O QUE ESTE ARQUIVO CONTÉM', 'h'),
    ('Resumo — os indicadores principais do estudo em uma tela.', 'p'),
    ('Cobertura — os 48 setores do estudo e quantas visitas cada um enviou.', 'p'),
    ('Tempo por canal — quanto cada bloco de atividade consome, em minutos por visita.', 'p'),
    ('Classificação do tempo — o tempo separado entre execução, apoio, deslocamento e espera.', 'p'),
    ('Movimentos — o detalhe completo, movimento a movimento. É a base das demais abas.', 'p'),
    ('Variabilidade — a dispersão do tempo de visita dentro de cada canal.', 'p'),
    ('Visitas — uma linha por visita medida.', 'p'),
    ('Deslocamento — os trajetos até a loja medidos até agora.', 'p'),
    ('', ''),
    ('COMO OS NÚMEROS FORAM CALCULADOS', 'h'),
    ('As atividades se sobrepõem no tempo: o promotor atende duas ou três categorias na mesma ida', 'p'),
    ('à gôndola e registra a mesma janela de horário para cada uma. Somar as durações declaradas', 'p'),
    ('infla o total em cerca de 2,4 vezes e faria os percentuais passarem de 100%.', 'p'),
    ('', ''),
    ('Por isso o tempo é alocado minuto a minuto: percorre-se a linha do tempo da visita e, em cada', 'p'),
    ('minuto, o tempo é dividido igualmente entre as atividades ativas naquele minuto. Dois movimentos', 'p'),
    ('feitos em paralelo levam metade do crédito cada. Assim a soma de todos os movimentos é', 'p'),
    ('exatamente o tempo real dentro da loja, e os percentuais fecham em 100%.', 'p'),
    ('', ''),
    ('O QUE FOI EXCLUÍDO DA BASE', 'h'),
    ('Atividades com término anterior ao início (inversão de digitação) foram tratadas como inválidas.', 'p'),
    ('Atividades com mais de 4 horas de duração foram descartadas como erro de digitação.', 'p'),
    ('Envios repetidos da mesma visita foram consolidados, mantendo-se a versão mais completa.', 'p'),
    ('', ''),
    ('LIMITES DESTA LEITURA — LEIA ANTES DE DECIDIR', 'h'),
    ('1. O tempo medido aqui é o tempo DENTRO da loja. O trajeto entre lojas tem apenas '
     '{} medições'.format(desl['n']), 'p'),
    ('   e não entra nas projeções. Uma rota real é mais longa do que a projeção sugere.', 'p'),
    ('2. CLUB, LASA e Perfumaria têm poucas visitas. São indício, não média: uma visita atípica', 'p'),
    ('   move o número inteiro. Estão marcados como amostra reduzida nas abas.', 'p'),
    ('3. A projeção "lojas por 8h" é aritmética simples (480 min ÷ tempo médio de visita). Não é', 'p'),
    ('   meta operacional: ignora deslocamento, pausas e imprevistos.', 'p'),
    ('4. Cada promotor mediu a própria jornada. Não houve cronometragem por observador externo.', 'p'),
]
r = 1
for txt, k in LINHAS:
    cel = ws.cell(row=r, column=1, value=txt)
    if k == 't':
        cel.font = h1
        ws.row_dimensions[r].height = 22
    elif k == 's':
        cel.font = it
    elif k == 'h':
        cel.font = h2
        ws.row_dimensions[r].height = 20
    else:
        cel.font = nm
    r += 1

# ------------------------------------------------------------- Movimentos
# Escrita primeiro porque as outras abas apontam para ela.
mv = mov.sort_values(['canal', 'bloco', 'min_equiv'],
                     ascending=[True, True, False]).reset_index(drop=True)
ws = wb.create_sheet('Movimentos')
ws.cell(row=1, column=1, value='Detalhe completo — todos os movimentos medidos, por canal').font = h1
ws.cell(row=2, column=1, value='Min equivalente = minutos alocados ao movimento, já descontada a '
        'sobreposição entre atividades simultâneas.').font = it
cabecalho(ws, 4,
          ['Canal', 'Bloco', 'Movimento', 'Classificação', 'Min equivalente (total)',
           'Visitas do canal', 'Min por visita', '% do tempo de loja'],
          [12, 18, 48, 22, 15, 12, 12, 14])
NM = 4 + len(mv)
for i, x in mv.iterrows():
    R = 5 + i
    for j, val in enumerate([x.canal, x.bloco, x.movimento, x.classe], 1):
        ws.cell(row=R, column=j, value=val).font = nm
    ws.cell(row=R, column=5, value=round(float(x.min_equiv), 1)).font = nm
    ws.cell(row=R, column=6, value=int(nvis[x.canal])).font = nm
    ws.cell(row=R, column=7, value='=E{0}/F{0}'.format(R)).font = nm
    ws.cell(row=R, column=8,
            value='=E{0}/SUMIFS($E$5:$E${1},$A$5:$A${1},$A{0})'.format(R, NM)).font = nm
    for c in (5, 7):
        ws.cell(row=R, column=c).number_format = '0.0'
    ws.cell(row=R, column=8).number_format = '0.0%'
    zebrar(ws, R, 8, i)

# ----------------------------------------------------------------- Resumo
ws = wb.create_sheet('Resumo', 1)
for col, w in zip('ABC', (46, 16, 56)):
    ws.column_dimensions[col].width = w
ws.cell(row=1, column=1, value='Resumo executivo').font = h1
ws.cell(row=2, column=1,
        value='Coleta de {} a {} · {} envios consolidados em {} visitas'.format(
            *meta['periodo'], meta['linhas_brutas'], len(vis))).font = it

K = [('COBERTURA', 'h', ''),
     ('Setores participantes', '{} de {}'.format(meta['setores'], meta['setores_total']),
      'Todos os setores do estudo enviaram ao menos uma visita'),
     ('Promotores distintos', vis.promotor.nunique(), 'Cada promotor mediu a própria jornada'),
     ('Visitas medidas', len(vis), 'Uma visita = uma passagem completa por uma loja'),
     ('Canais cobertos', vis.canal.nunique(), 'DPP, C&C, NMR, GMR, HFS, CLUB, LASA e Perfumaria'),
     ('', '', ''),
     ('TEMPO DENTRO DA LOJA', 'h', ''),
     ('Tempo mediano de visita', '{:.0f} min'.format(vis.tempo_loja_min.median()),
      'Mediana entre todas as visitas, todos os canais'),
     ('Visita mais curta', '{:.0f} min'.format(vis.tempo_loja_min.min()), 'Canal DPP'),
     ('Visita mais longa', '{:.0f} min'.format(vis.tempo_loja_min.max()), 'Canal GMR'),
     ('', '', ''),
     ('COMO O TEMPO É GASTO DENTRO DA LOJA', 'h', '')]
for k, v in extra['classes'].items():
    K.append((k, '{:.1f}%'.format(v['pct']),
              'Equivale a {}h{:02d} de uma jornada de 8 horas'.format(v['min_8h'] // 60, v['min_8h'] % 60)))
K += [('', '', ''),
      ('DESLOCAMENTO ATÉ A LOJA (amostra reduzida)', 'h', ''),
      ('Trajetos medidos', desl['n'],
       'Por {} promotores — não entra nas projeções de jornada'.format(desl['promotores']))]
for origem, rot in (('Casa / base', 'De casa até a primeira loja'), ('Loja anterior', 'Entre duas lojas')):
    d = desl['por_origem'].get(origem)
    if d:
        K.append((rot, '{} min'.format(d['mediana']), '{} medições (mediana)'.format(d['n'])))

r = 4
for a, b, c in K:
    if b == 'h':
        ws.cell(row=r, column=1, value=a).font = h2
        ws.row_dimensions[r].height = 20
    elif a:
        ws.cell(row=r, column=1, value=a).font = nm
        cel = ws.cell(row=r, column=2, value=b)
        cel.font = bd
        cel.alignment = Alignment(horizontal='right')
        ws.cell(row=r, column=3, value=c).font = it
        for cc in range(1, 4):
            ws.cell(row=r, column=cc).border = thin
    r += 1

# -------------------------------------------------------------- Cobertura
ws = wb.create_sheet('Cobertura', 2)
ws.cell(row=1, column=1, value='Cobertura do estudo').font = h1
ws.cell(row=2, column=1, value='Os 48 setores selecionados e quantas visitas cada um enviou.').font = it
cabecalho(ws, 4, ['Setor', 'Promotor', 'Executivo', 'Canais', 'Lojas assignadas',
                  'Visitas enviadas', 'Status'], [10, 36, 14, 14, 15, 14, 13])
cs = cob.sort_values(['Executivo', 'Setor']).reset_index(drop=True)
for i, x in cs.iterrows():
    R = 5 + i
    vals = [x.Setor, x.Promotor if pd.notna(x.Promotor) else '—', x.Executivo, x.Canais,
            int(x['Lojas assignadas']), int(x['Visitas enviadas']), x.Status]
    for j, val in enumerate(vals, 1):
        ws.cell(row=R, column=j, value=val).font = nm
    ws.cell(row=R, column=7).font = Font(
        name=F, size=10, bold=True,
        color='2F6B4F' if x.Status == 'Respondeu' else 'A8452A')
    zebrar(ws, R, 7, i)

# -------------------------------------------------------- Tempo por canal
ws = wb.create_sheet('Tempo por canal', 3)
ws.cell(row=1, column=1, value='Tempo por bloco de atividade, por canal').font = h1
ws.cell(row=2, column=1, value='Minutos médios por visita. Sai por fórmula da aba Movimentos.').font = it
cabecalho(ws, 4, ['Canal', 'Visitas', 'Min por visita'] + BLOCOS +
          ['Projeção lojas/8h', 'Lojas/dia declarado', 'Amostra'],
          [12, 9, 13] + [13] * 8 + [15, 16, 16])
vd = vis.copy()
vd.loc[vd.lojas_no_dia > 12, 'lojas_no_dia'] = pd.NA
for i, ch in enumerate(ordc):
    R = 5 + i
    ws.cell(row=R, column=1, value=ch).font = bd
    ws.cell(row=R, column=2, value=int(nvis[ch])).font = nm
    ws.cell(row=R, column=3, value='=SUM(D{0}:K{0})'.format(R)).font = bd
    for j in range(len(BLOCOS)):
        col = get_column_letter(4 + j)
        ws.cell(row=R, column=4 + j, value=(
            '=SUMIFS(Movimentos!$E$5:$E${1},Movimentos!$A$5:$A${1},$A{0},'
            'Movimentos!$B$5:$B${1},{2}$4)/$B{0}').format(R, NM, col)).font = nm
    ws.cell(row=R, column=12, value='=480/$C{0}'.format(R)).font = nm
    ld = vd[vd.canal == ch].lojas_no_dia.median()
    ws.cell(row=R, column=13, value=None if pd.isna(ld) else float(ld)).font = nm
    ws.cell(row=R, column=14, value='Reduzida' if ch in POUCAS else 'Adequada').font = nm
    for c in range(3, 14):
        ws.cell(row=R, column=c).number_format = '0.0'
    zebrar(ws, R, 14, i)
ws.cell(row=6 + len(ordc), column=1,
        value='Amostra reduzida (menos de 8 visitas): tratar como indício, não como média.').font = it

# ------------------------------------------------- Classificação do tempo
ws = wb.create_sheet('Classificação do tempo', 4)
ws.cell(row=1, column=1, value='Classificação do tempo em loja').font = h1
ws.cell(row=2, column=1, value='Execução = trabalho na gôndola (abastecer, precificar, limpar, montar). '
        'Deslocamento e busca = andar e procurar. Espera = portaria, cadastro, EPI.').font = it
cabecalho(ws, 4, ['Canal', 'Min por visita'] + CLASSES + ['% Execução', 'Amostra'],
          [12, 13, 20, 18, 20, 18, 12, 14])
for i, ch in enumerate(ordc):
    R = 5 + i
    ws.cell(row=R, column=1, value=ch).font = bd
    ws.cell(row=R, column=2, value='=SUM(C{0}:F{0})'.format(R)).font = bd
    for j in range(len(CLASSES)):
        col = get_column_letter(3 + j)
        ws.cell(row=R, column=3 + j, value=(
            '=SUMIFS(Movimentos!$E$5:$E${1},Movimentos!$A$5:$A${1},$A{0},'
            'Movimentos!$D$5:$D${1},{2}$4)/{3}').format(R, NM, col, int(nvis[ch]))).font = nm
    ws.cell(row=R, column=7, value='=C{0}/$B{0}'.format(R)).font = bd
    ws.cell(row=R, column=8, value='Reduzida' if ch in POUCAS else 'Adequada').font = nm
    for c in range(2, 7):
        ws.cell(row=R, column=c).number_format = '0.0'
    ws.cell(row=R, column=7).number_format = '0.0%'
    zebrar(ws, R, 8, i)

# --------------------------------------------------------- Variabilidade
ws = wb.create_sheet('Variabilidade', 6)
ws.cell(row=1, column=1, value='Variabilidade do tempo de visita').font = h1
ws.cell(row=2, column=1, value='CV = desvio padrão ÷ média. Quanto maior, menos padronizada está a '
        'operação naquele canal — e maior o ganho possível com padronização.').font = it
cabecalho(ws, 4, ['Canal', 'Visitas', 'Mínimo (min)', 'Mediana (min)', 'Média (min)',
                  'Máximo (min)', 'Desvio padrão', 'CV', 'Amplitude (máx÷mín)'],
          [12, 9, 13, 13, 13, 13, 13, 10, 17])
d = vis.groupby('canal').tempo_loja_min.agg(['size', 'min', 'median', 'mean', 'max', 'std'])
for i, ch in enumerate(ordc):
    R = 5 + i
    x = d.loc[ch]
    ws.cell(row=R, column=1, value=ch).font = bd
    for j, val in enumerate([int(x['size']), float(x['min']), float(x['median']),
                             float(x['mean']), float(x['max']),
                             float(x['std']) if pd.notna(x['std']) else None], 2):
        ws.cell(row=R, column=j, value=val).font = nm
    ws.cell(row=R, column=8, value='=G{0}/E{0}'.format(R)).font = nm
    ws.cell(row=R, column=8).number_format = '0%'
    ws.cell(row=R, column=9, value='=F{0}/C{0}'.format(R)).font = nm
    ws.cell(row=R, column=9).number_format = '0.0"x"'
    for c in range(3, 8):
        ws.cell(row=R, column=c).number_format = '0'
    zebrar(ws, R, 9, i)

# --------------------------------------------------------------- Visitas
ws = wb.create_sheet('Visitas', 7)
ws.cell(row=1, column=1, value='Base de visitas').font = h1
ws.cell(row=2, column=1, value='Uma linha por visita medida. "Movimentos registrados" = quantas '
        'atividades tiveram horário válido naquela visita.').font = it
cabecalho(ws, 4, ['Promotor', 'Canal', 'Rede', 'Loja', 'Tempo de loja (min)', 'Tempo de loja (h)',
                  'Lojas no dia (declarado)', 'Movimentos registrados'],
          [32, 10, 26, 40, 15, 14, 17, 17])
vv = vis.sort_values(['canal', 'tempo_loja_min'], ascending=[True, False]).reset_index(drop=True)
for i, x in vv.iterrows():
    R = 5 + i
    for j, val in enumerate([x.promotor, x.canal, x.rede, x.loja, float(x.tempo_loja_min)], 1):
        ws.cell(row=R, column=j, value=val).font = nm
    ws.cell(row=R, column=5).number_format = '0'
    ws.cell(row=R, column=6, value='=E{0}/60'.format(R)).font = nm
    ws.cell(row=R, column=6).number_format = '0.00'
    ws.cell(row=R, column=7, value=None if pd.isna(x.lojas_no_dia) else float(x.lojas_no_dia)).font = nm
    ws.cell(row=R, column=8, value=int(x.movimentos)).font = nm
    zebrar(ws, R, 8, i)

# ---------------------------------------------------------- Deslocamento
ws = wb.create_sheet('Deslocamento', 8)
ws.cell(row=1, column=1, value='Deslocamento até a loja').font = h1
ws.cell(row=2, column=1, value='Bloco incluído no formulário em 15/08, depois do início da coleta. '
        'Amostra ainda reduzida — não entra nas projeções de jornada.').font = it
cabecalho(ws, 4, ['Origem do trajeto', 'Medições', 'Mediana (min)'], [30, 12, 15])
r = 5
for k, v in desl['por_origem'].items():
    ws.cell(row=r, column=1, value=k).font = nm
    ws.cell(row=r, column=2, value=v['n']).font = nm
    ws.cell(row=r, column=3, value=v['mediana']).font = bd
    r += 1
for txt in ('Total: {} trajetos medidos por {} promotores. Mediana geral de {} min.'.format(
                desl['n'], desl['promotores'], desl['mediana']),
            'O trajeto de casa até a primeira loja é o mais longo; entre duas lojas o tempo cai bastante.',
            'Enquanto a amostra não crescer, considere a rota real mais longa do que a projeção "lojas por 8h".'):
    r += 1
    ws.cell(row=r, column=1, value=txt).font = it

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

SAIDA = 'analise/entregaveis/Tempos e Movimentos - Base de Dados.xlsx'
wb.save(SAIDA)
print('salvo:', SAIDA)
print('abas:', ', '.join(s.title for s in wb.worksheets))
